from openpyxl import load_workbook
# get the workbook
file = load_workbook("inventory.xlsx")
# open the active sheet- we can also do:
# sheet = file["Sheet1"]
active_sheet = file.active

# Calculates {"CompanyA: numofproducts", "CompanyB: numofproducts"}
def calculate_products_by_company(worksheet):
    # dict to store count of products by each company
    company_products = dict()
    # reading values from each cell of just 4th column, starting from 2nd row
    for company in worksheet.iter_rows(min_row=2, min_col=4, max_col=4, values_only=True):
        # iter_rows returns a tuple like ('Company Name',) hence extracting company name at [0]
        company_name = company[0]
        if company_products.get(company_name) is None:
            company_products[company_name] = 1
        else:
            company_products[company_name] += 1
    # print the output dict values
    for company_name, product_count in company_products.items():
        print(f"{company_name}: {product_count}")

# Calculates {"CompanyA: inventory_value", "CompanyB: total_inventory_value"}
def total_inventory_value_by_company(worksheet):
    # dict to store count of inventory value by each company
    company_inventory_value = dict()
    # reading values from each cell of 2, 3 and 4th column, starting from 2nd row
    for each_row in worksheet.iter_rows(min_row=2, min_col=2, max_col=4, values_only=True):
        # iter_rows returns a tuple like ('Inventory Value', 'Price', 'Company Name',)
        inventory = each_row[0]
        price = each_row[1]
        company_name = each_row[2]
        inventory_value = inventory * price
        if company_inventory_value.get(company_name) is None:
            company_inventory_value[company_name] = inventory_value
        else:
            company_inventory_value[company_name] += inventory_value
    for company_name, inventory_value in company_inventory_value.items():
        print(f"{company_name}: {inventory_value}")

# Calclulates [(ProductNo, Inventory), (ProductNo, Inventory)]
def products_with_less_than_10_inventory(worksheet):
    # list to store products
    product_data = list(tuple())
    # reading values from each cell of each column, starting from 2nd row
    for each_row in worksheet.iter_rows(min_row=2, min_col=1, max_col=4, values_only=True):
        # iter_rows returns a tuple like ('Product No', 'Inventory')
        inventory = each_row[1]
        if inventory < 10:
            product_data.append(each_row)
    for each_row in product_data:
        print(each_row)

# Writes inventory value for each row in a new column (currently saves as a new file)
def write_inventory_value_for_each_product(worksheet):
    # reading values from each cell of 2, 3 column, starting from 2nd row
    for each_row in worksheet.iter_rows(min_row=2, min_col=1, max_col=3, values_only=True):
        # iter_rows returns a tuple like ('Product No', 'Inventory', 'Price',)
        row_number = int(each_row[0]) + 1
        inventory = each_row[1]
        price = each_row[2]
        inventory_value = inventory * price
        worksheet.cell(row=row_number, column=5, value=inventory_value)
    # save file with a new name
    file.save("edited_xl.xlsx")

# calculate_products_by_company(active_sheet)
# total_inventory_value_by_company(active_sheet)
# products_with_less_than_10_inventory(active_sheet)
# write_inventory_value_for_each_product(active_sheet)

